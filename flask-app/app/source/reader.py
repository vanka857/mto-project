import openpyxl
import re
import pandas as pd


def extract_inventory_details(sheet):
    inventory_descriptions = []
    in_description = False
    current_items = []
    description_details = {}

    # Перебор всех строк листа
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        cell_value = row[26] if len(row) > 26 else None  # Смотрим в столбец AA (R1C27)
        
        # Поиск начала описи
        if cell_value and 'ИНВЕНТАРИЗАЦИОННАЯ ОПИСЬ' in cell_value:
            # Если мы уже внутри описи, сохраняем предыдущую опись и начинаем новую
            if in_description:
                description_details['Items'] = current_items
                inventory_descriptions.append(description_details)
                current_items = []
            
            in_description = True
            description_details = {
                'Number': re.search(r'\d+-\d+$', cell_value).group(),
                'Reffered date': sheet.cell(row=i+3, column=46).value,  # Дата описи на 3 строки ниже
                'Responsible Person': sheet.cell(row=i+28, column=35).value  # Ответственное лицо
            }
        
        # Сбор данных об элементах, если мы внутри описи
        if in_description and row[0] is not None and isinstance(row[0], int):  # Предполагаем, что ID элемента — это число
            item_details = {
                'Number': row[0] if len(row) > 0 else None,
                'Name': row[3] if len(row) > 3 else None,
                'Inventory Number': str(row[7]).strip() if len(row) > 7 else None,
                'Measure': row[10] if len(row) > 10 else None,
                'Volume': float(row[15]) if len(row) > 15 else None,
                'Price': float(row[12]) if len(row) > 12 else None
            }
            current_items.append(item_details)
    
    # Добавляем последнюю опись, если такая имеется
    if in_description and current_items:
        description_details['Items'] = current_items
        inventory_descriptions.append(description_details)

    return inventory_descriptions

def process_inventory_file(file_path):
    workbook = openpyxl.load_workbook(file_path)
    results = []

    # Обработка всех листов в книге
    for sheet in workbook.worksheets:
        results.extend(extract_inventory_details(sheet))

    return results

def read_excel_file(filepath):
    try:
        inventories = process_inventory_file(filepath)  # Предполагается, что эта функция возвращает структуру данных как показано

        inventory_list = []
        for inventory in inventories:
            number = inventory['Number']
            owner = inventory['Responsible Person']
            date = inventory['Reffered date']
            inventory_data = pd.DataFrame(inventory['Items'])

            # Сохраняем данные в более подходящем для работы формате
            items_data = inventory_data.to_dict(orient='records')  # Преобразуем DataFrame в список словарей

            # Сохранение результатов в список
            inventory_list.append({
                'number': number,
                'owner': owner,
                'date': date,
                'items': items_data  # сохраняем исходные данные, а не HTML
            })
        
        return 'Success', 'File successfully read', inventory_list
    except Exception as e:
        return 'Failed', str(e), []

# # ------------ ДЛЯ ТЕСТОВЫХ ЗАПУСКОВ
# # Путь к файлу
# file_path = '/Users/vankudr/Documents/ВК/project/legacy/table2.xlsx'

# # Выполнение функции и вывод результатов
# inventories = process_inventory_file(file_path)
# for inventory in inventories:
#     print(f"Номер описи: {inventory['Number']}")
#     print(f"Дата описи: {inventory['Reffered date']}")
#     print(f"Ответственное лицо: {inventory['Responsible Person']}")

#     inventory_data = pd.DataFrame(inventory['Items'])
#     print(f"Итого по описи записей: {len(inventory_data)}")
#     print(f"Итого по описи (штук): {int(inventory_data['Volume'].sum())}")
#     print(f"Итого по описи (стоимость, руб): {round(sum(inventory_data['Volume'] * inventory_data['Price']), 2)}")
#     # for item in inventory['Items']:
#     #     print(item)
